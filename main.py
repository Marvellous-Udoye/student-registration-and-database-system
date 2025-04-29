import os
from tkinter import *
from tkinter import filedialog, messagebox
from tkinter.ttk import Combobox
from PIL import Image, ImageTk
from datetime import date
import openpyxl
import pathlib
from openpyxl import Workbook


class StudentRegistrationSystem:
    def __init__(self, root):
        self.root = root
        self.root.title("Student Registration System")
        self.root.geometry("1250x700+50+20")
        self.background = "#06283D"
        self.framebg = "#EDEDED"
        self.framefg = "#06283D"
        self.root.config(bg=self.background)

        self.setup_ui()
        self.check_and_create_excel_file()

    def setup_ui(self):
        Label(self.root, text="Email: Team-1_Wed_2-4@gmail.com", width=10, height=3, bg="#f0687c", anchor='e').pack(side=TOP, fill=X)
        Label(self.root, text="STUDENT REGISTRATION AND DATABASE SYSTEM", width=10, height=2, bg="#c36464", fg='#fff', font='arial 20 bold').pack(side=TOP, fill=X)

        self.setup_registration_and_date()
        self.setup_personal_info()
        self.setup_general_info()
        self.setup_image_frame()
        self.setup_buttons()

    def setup_registration_and_date(self):
        Label(self.root, text="Date:", font="arial 13", fg=self.framebg, bg=self.background).place(x=750, y=150)

        self.date = StringVar()

        today = date.today()
        d1 = today.strftime("%d/%m/%Y")
        Entry(self.root, textvariable=self.date, width=15, font="arial 10").place(x=820, y=150)
        self.date.set(d1)

    def setup_personal_info(self):
        obj = LabelFrame(self.root, text="Personal Information", font=20, bd=2, width=900, bg=self.framebg, fg=self.framefg, height=220, relief=GROOVE)
        obj.place(x=30, y=200)

        Label(obj, text="Full Name:", font="arial 13", bg=self.framebg, fg=self.framefg).place(x=30, y=50)
        Label(obj, text="Matric. no:", font="arial 13", bg=self.framebg, fg=self.framefg).place(x=30, y=100)
        Label(obj, text="Gender:", font="arial 13", bg=self.framebg, fg=self.framefg).place(x=30, y=150)

        Label(obj, text="Level:", font="arial 13", bg=self.framebg, fg=self.framefg).place(x=500, y=50)
        Label(obj, text="Religion:", font="arial 13", bg=self.framebg, fg=self.framefg).place(x=500, y=100)
        Label(obj, text="Date of Birth:", font="arial 13", bg=self.framebg, fg=self.framefg).place(x=500, y=150)

        self.name = StringVar()
        Entry(obj, textvariable=self.name, width=20, font="arial 10").place(x=160, y=50)

        self.matriculation = StringVar()
        Entry(obj, textvariable=self.matriculation, width=20, font="arial 10").place(x=160, y=100)

        self.radio = IntVar()
        Radiobutton(obj, text="Male", variable=self.radio, value=1, bg=self.framebg, fg=self.framefg, command=self.selection).place(x=150, y=150)
        Radiobutton(obj, text="Female", variable=self.radio, value=2, bg=self.framebg, fg=self.framefg, command=self.selection).place(x=200, y=150)

        self.religion = Combobox(obj, values=['Christianity', 'Islam', 'Others'], font="Roboto 10", width=17, state="r")
        self.religion.place(x=630, y=100)
        self.religion.set("Select Religion")

        self.dob = StringVar()
        today = date.today()
        d2 = today.strftime("%d/%m/%Y")
        Entry(obj, textvariable=self.dob, width=20, font="arial 10").place(x=630, y=150)
        self.dob.set(d2)

        self.level = Combobox(obj, values=['100', '200', '300', '400', '500'], font="Roboto 10", width=17, state="r")
        self.level.place(x=630, y=50)
        self.level.set("Select Level")

    def setup_general_info(self):
        obj2 = LabelFrame(self.root, text="General Information", font=20, bd=2, width=900, bg=self.framebg, fg=self.framefg, height=220, relief=GROOVE)
        obj2.place(x=30, y=450)

        Label(obj2, text="College:", font="arial 13", bg=self.framebg, fg=self.framefg).place(x=30, y=50)
        Label(obj2, text="Course:", font="arial 13", bg=self.framebg, fg=self.framefg).place(x=30, y=100)

        self.college = Combobox(obj2, values=['COLENG', 'COLNAS', 'COLENVS', 'COLMANS', 'COLFAST'], font="Roboto 10", width=17, state="r")
        self.college.place(x=120, y=50)
        self.college.set("Select College")

        self.department = StringVar()
        Entry(obj2, textvariable=self.department, width=20, font="arial 10").place(x=120, y=100)

        Label(obj2, text="Email Address:", font="arial 13", bg=self.framebg, fg=self.framefg).place(x=500, y=50)
        Label(obj2, text="Phone Number:", font="arial 13", bg=self.framebg, fg=self.framefg).place(x=500, y=100)

        self.email = StringVar()
        Entry(obj2, textvariable=self.email, width=20, font="arial 10").place(x=630, y=50)

        self.phone = StringVar()
        Entry(obj2, textvariable=self.phone, width=20, font="arial 10").place(x=630, y=100)

    def setup_image_frame(self):
        f = Frame(self.root, bd=3, bg="black", width=200, height=200, relief=GROOVE)
        f.place(x=996, y=146)

        self.img_path = "images/images (1).png"
        self.img = PhotoImage(file=self.img_path)
        self.lbl = Label(self.root, bg="black", image=self.img, width=190, height=190)
        self.lbl.place(x=1000, y=150)

    def setup_buttons(self):
        Button(self.root, text="Upload", width=19, height=2, font="arial 12 bold", bg="lightblue", command=self.showimage).place(x=1000, y=370)
        self.saveButton = Button(self.root, text="Save", width=19, height=2, font="arial 12 bold", bg="lightgreen", command=self.save)
        self.saveButton.place(x=1000, y=450)
        Button(self.root, text="Reset", width=19, height=2, font="arial 12 bold", bg="lightpink", command=self.reset).place(x=1000, y=530)
        Button(self.root, text="Exit", width=19, height=2, font="arial 12 bold", bg="grey", command=self.exit).place(x=1000, y=620)

    def check_and_create_excel_file(self):
        file = pathlib.Path('Student_data.xlsx')
        if not file.exists():
            file = Workbook()
            sheet = file.active
            sheet['A1'] = "Date"
            sheet['B1'] = "Matric. No."
            sheet['C1'] = "Full Name"
            sheet['D1'] = "Gender"
            sheet['E1'] = "Level"
            sheet['F1'] = "College"
            sheet['G1'] = "Department"
            sheet['H1'] = "Date of Birth"
            sheet['I1'] = "Religion"
            sheet['J1'] = "Email Address"
            sheet['K1'] = "Phone Number"

            file.save('Student_data.xlsx')

    def showimage(self):
        filename = filedialog.askopenfilename(initialdir=os.getcwd(), title="Select image file",
                                              filetypes=(("JPG File", "*.jpg"), ("PNG File", "*.png"), ("All files", "*.*")))
        if filename:
            img = Image.open(filename)
            resized_image = img.resize((190, 190), Image.Resampling.LANCZOS)
            photo_2 = ImageTk.PhotoImage(resized_image)
            self.lbl.config(image=photo_2)
            self.lbl.image = photo_2
            self.img_path = filename

    def selection(self):
        value = self.radio.get()
        self.gender = "Male" if value == 1 else "Female"


    def reset(self):
        self.name.set('')
        self.matriculation.set('')
        self.religion.set('')
        self.college.set('')
        self.email.set('')
        self.department.set('')
        self.phone.set('')
        self.level.set('Select Level')

        self.saveButton.config(state='normal')

        img1 = PhotoImage(file='images/images (1).png')
        self.lbl.config(image=img1)
        self.lbl.image = img1

        self.img_path = "images/gallery(2).png"

    def save(self):
        G1 = None
        N1 = self.name.get()
        L1 = self.level.get()
        try:
            G1 = self.gender
        except:
            messagebox.showerror("Error", "Select Gender!")
            return
        M1 = self.matriculation.get()
        D1 = self.date.get()
        Rel = self.religion.get()
        D2 = self.dob.get()
        C1 = self.college.get()
        mail = self.email.get()
        dept = self.department.get()
        P1 = self.phone.get()

        if N1 == "" or L1 == "Select Level" or Rel == "" or M1 == "" or D2 == "" or C1 == "" or mail == "" or dept == "" or P1 == "":
            messagebox.showerror("Error", "Few Data is Missing!")
        else:
            file = openpyxl.load_workbook('Student_data.xlsx')
            sheet = file.active
            sheet.cell(column=1, row=sheet.max_row + 1, value=D1)
            sheet.cell(column=2, row=sheet.max_row, value=M1)
            sheet.cell(column=3, row=sheet.max_row, value=N1)
            sheet.cell(column=4, row=sheet.max_row, value=G1)
            sheet.cell(column=5, row=sheet.max_row, value=L1)
            sheet.cell(column=6, row=sheet.max_row, value=C1)
            sheet.cell(column=7, row=sheet.max_row, value=dept)
            sheet.cell(column=8, row=sheet.max_row, value=D2)
            sheet.cell(column=9, row=sheet.max_row, value=Rel)
            sheet.cell(column=10, row=sheet.max_row, value=mail)
            sheet.cell(column=11, row=sheet.max_row, value=P1)

            try:
                img = Image.open(self.img_path)
                img.save("Student_images/" + str(M1) + ".jpg")
                messagebox.showinfo("Info", "Form Filled Successfully")
                self.reset()
                file.save(r'Student_data.xlsx')
            except Exception as e:
                messagebox.showinfo("Info", "Profile Picture is not Available!!")
                # print(e)

    def exit(self):
        self.root.destroy()


if __name__ == "__main__":
    root = Tk()
    app = StudentRegistrationSystem(root)
    root.mainloop()
